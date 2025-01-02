import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill
from helper import *
from tkinter import *

# pdf_file = "ts-des24.pdf"
# excel_file = "ts-des24 - Copy.xlsx"
# start_row_excel = 13

# workbook = openpyxl.load_workbook(excel_file)
# sheet = workbook.active

def writeValue(start_row_excel, pdf_file, excel_file, button):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        date_column = 'C'
        in_column = 'D'
        out_column = 'E'
        
        date_pdf_column_index = 1
        in_pdf_column_index = 2
        out_pdf_column_index = 3

        with pdfplumber.open(pdf_file) as pdf:
            tables = pdf.pages[0].extract_tables()
            
            for row in tables[0]:
                if len(row) > out_pdf_column_index:
                    date_val = row[date_pdf_column_index]
                    
                    if date_val == "Date":
                        continue
                    if isWeekend(date_val):
                        continue
                    
                    date_val = date_val.replace("-", "/")
                    in_val = row[in_pdf_column_index]
                    out_val = row[out_pdf_column_index]
                    
                    date_cell = sheet[f"{date_column}{start_row_excel}"]
                    date_cell.value = date_val
                    date_cell.fill = PatternFill()
                    
                    sheet[f"{in_column}{start_row_excel}"] = in_val
                    sheet[f"{out_column}{start_row_excel}"] = out_val
                    
                    start_row_excel += 1

        workbook.save(excel_file)
        button.config(state="enable")
        print(f"Data has been successfully written to {excel_file}")
        return "successfully"

    except Exception as e:
        button.config(state="enable")
        print(f"An error occurred: {e}")
        return f"An error occurred: {e}"


# writeValue(start_row_excel)